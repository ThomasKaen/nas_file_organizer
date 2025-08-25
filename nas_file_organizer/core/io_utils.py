from __future__ import annotations
import io
from pathlib import Path
from datetime import datetime
from typing import Iterable, Sequence
import chardet
import fitz                         # PyMuPDF
from PIL import Image
import pytesseract
from docx import Document
from openpyxl import load_workbook
from .cache import get_text as cache_get, set_text as cache_set


def list_files(folder: Path) -> Iterable[Path]:
    for p in folder.rglob("*"):
        if p.is_file():
            yield p

def read_text_any(
    path: Path,
    ocr_lang: str = "eng",
    skip_large_mb: int = 50,
    page_window_first: int = 2,
    page_window_last: int = 1,
    ocr_on_empty_text: bool = True,
) -> str:
    cached = cache_get(path)
    if cached is not None:
        return cached
    try:
        if path.stat().st_size > skip_large_mb * 1024 * 1024:
            return ""
    except Exception:
        pass
    ext = path.suffix.lower()

    if ext == ".pdf":
        text = _pdf_text(path, ocr_lang, page_window_first, page_window_last, ocr_on_empty_text)
        if text.strip():
            cache_set(path, text)
        return text

    if ext in {".txt", ".log", ".md"}:
        text = _txt_text(path)
        if text.strip():
            cache_set(path, text)
        return text

    if ext == ".docx":
        text = _docx_text(path)
        if text.strip():
            cache_set(path, text)
        return text

    if ext == ".xlsx":
        text = _xlsx_text(path)
        if text.strip():
            cache_set(path, text)
        return text

    if ext in {".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"}:
        text = _image_ocr(path, ocr_lang)
        if text.strip():
            cache_set(path, text)
        return text

    return ""

def _pdf_text(path: Path, ocr_lang: str, first_n: int, last_n: int, ocr_on_empty: bool) -> str:
    try:
        doc = fitz.open(path)
        n = len(doc)
        idxs: list[int] = list(range(min(first_n, n)))
        # add last pages if requested and not overlapping
        for i in range(max(0, n - last_n), n):
            if i not in idxs:
                idxs.append(i)

        parts: list[str] = []
        for i in idxs:
            page = doc[i]
            t = page.get_text("text") or ""
            if (not t.strip()) and ocr_on_empty:
                # low-DPI probe OCR (fast). If you want, bump to 300 later on poor results.
                pix = page.get_pixmap(dpi=200)
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                t = pytesseract.image_to_string(img, lang=ocr_lang)
            parts.append(t)
        return "\n".join(parts)
    except Exception:
        return ""

def _image_ocr(path: Path, ocr_lang: str) -> str:
    try:
        img = Image.open(path)
        return pytesseract.image_to_string(img, lang=ocr_lang)
    except Exception:
        return ""

def _txt_text(path: Path) -> str:
    raw = path.read_bytes()
    enc = chardet.detect(raw).get("encoding") or "utf-8"
    try:
        return raw.decode(enc, errors="ignore")
    except Exception:
        return ""

def _docx_text(path: Path) -> str:
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""

def _xlsx_text(path: Path) -> str:
    try:
        wb = load_workbook(path, data_only=True)
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                out.append(" ".join("" if v is None else str(v) for v in row))
        return "\n".join(out)
    except Exception:
        return ""

def next_available(p: Path) -> Path:
    if not p.exists(): return p
    stem, suf = p.stem, p.suffix
    i = 2
    while True:
        cand = p.with_name(f"{stem}_{i}{suf}")
        if not cand.exists(): return cand
        i += 1

def render_template(tmpl: str, *, original: str, date: datetime, archive_root: Path, first_keyword: str | None = None) -> str:
    return (tmpl
            .replace("{original}", original)
            .replace("{year}", f"{date.year}")
            .replace("{month}", f"{date.month:02d}")
            .replace("{day}", f"{date.day:02d}")
            .replace("{date}", date.strftime("%Y-%m-%d"))
            .replace("{archive_root}", str(archive_root))
            .replace("{first_keyword}", first_keyword or ""))
